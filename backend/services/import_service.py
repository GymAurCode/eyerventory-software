from __future__ import annotations

import io
from typing import Any

from sqlalchemy.orm import Session

from backend.models.product import Product
from backend.schemas.product import ImportRowResult, ImportSummary

# Required columns in the Excel file (case-insensitive)
REQUIRED_COLS = {"sku", "name", "price", "quantity"}
OPTIONAL_COLS = {"category"}
ALL_COLS = REQUIRED_COLS | OPTIONAL_COLS

# Sample template rows
TEMPLATE_ROWS = [
    {"sku": "SKU-001", "name": "Sample Product A", "price": 99.99, "quantity": 50, "category": "Electronics"},
    {"sku": "SKU-002", "name": "Sample Product B", "price": 149.50, "quantity": 30, "category": "Accessories"},
]


def generate_template_xlsx() -> bytes:
    """Return bytes of a sample .xlsx template file."""
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["sku", "name", "price", "quantity", "category"]
        ws.append(headers)
        for row in TEMPLATE_ROWS:
            ws.append([row[h] for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")


def _read_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse Excel bytes into a list of row dicts (lowercase keys)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")
        # Normalize headers to lowercase stripped strings
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")


def _validate_row(row: dict, row_num: int) -> tuple[dict | None, str | None]:
    """
    Validate a single row. Returns (cleaned_data, error_message).
    cleaned_data is None if validation fails.
    """
    sku = str(row.get("sku") or "").strip()
    if not sku:
        return None, "sku is required"

    name = str(row.get("name") or "").strip()
    if not name:
        return None, "name is required"
    if len(name) < 2:
        return None, "name must be at least 2 characters"

    raw_price = row.get("price")
    try:
        price = float(raw_price)
        if price <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return None, f"price must be a positive number (got: {raw_price!r})"

    raw_qty = row.get("quantity")
    try:
        qty = int(float(raw_qty))
        if qty < 0:
            raise ValueError
    except (TypeError, ValueError):
        return None, f"quantity must be a non-negative integer (got: {raw_qty!r})"

    category = str(row.get("category") or "").strip() or None

    return {
        "sku": sku,
        "name": name,
        "cost_price": price,
        "stock": qty,
        "category": category,
    }, None


def import_products(db: Session, file_bytes: bytes) -> ImportSummary:
    """
    UPSERT products from an Excel file.
    - Match existing products by SKU.
    - Insert new ones, update existing ones.
    - Skip and log invalid rows.
    """
    raw_rows = _read_excel(file_bytes)

    results: list[ImportRowResult] = []
    inserted = updated = failed = 0

    for i, raw in enumerate(raw_rows, start=2):  # row 1 = header
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in raw.values()):
            continue

        cleaned, error = _validate_row(raw, i)
        sku = str(raw.get("sku") or "").strip()

        if error:
            failed += 1
            results.append(ImportRowResult(
                row=i, sku=sku or "?", name=str(raw.get("name") or ""),
                action="failed", error=error,
            ))
            continue

        try:
            existing = db.query(Product).filter(Product.sku == cleaned["sku"]).first()
            if existing:
                # UPDATE existing product
                existing.name = cleaned["name"]
                existing.cost_price = cleaned["cost_price"]
                existing.stock = cleaned["stock"]
                existing.category = cleaned["category"]
                db.flush()
                updated += 1
                results.append(ImportRowResult(
                    row=i, sku=cleaned["sku"], name=cleaned["name"], action="updated",
                ))
            else:
                # INSERT new product — check name uniqueness too
                name_conflict = db.query(Product).filter(Product.name == cleaned["name"]).first()
                if name_conflict:
                    # Assign the SKU to the existing product if it has none
                    if not name_conflict.sku:
                        name_conflict.sku = cleaned["sku"]
                        name_conflict.cost_price = cleaned["cost_price"]
                        name_conflict.stock = cleaned["stock"]
                        name_conflict.category = cleaned["category"]
                        db.flush()
                        updated += 1
                        results.append(ImportRowResult(
                            row=i, sku=cleaned["sku"], name=cleaned["name"],
                            action="updated", error=None,
                        ))
                    else:
                        failed += 1
                        results.append(ImportRowResult(
                            row=i, sku=cleaned["sku"], name=cleaned["name"],
                            action="failed",
                            error=f"Name '{cleaned['name']}' already exists with a different SKU",
                        ))
                    continue

                product = Product(
                    name=cleaned["name"],
                    sku=cleaned["sku"],
                    cost_price=cleaned["cost_price"],
                    stock=cleaned["stock"],
                    category=cleaned["category"],
                )
                db.add(product)
                db.flush()
                inserted += 1
                results.append(ImportRowResult(
                    row=i, sku=cleaned["sku"], name=cleaned["name"], action="inserted",
                ))
        except Exception as exc:
            db.rollback()
            failed += 1
            results.append(ImportRowResult(
                row=i, sku=cleaned.get("sku", "?"), name=cleaned.get("name", ""),
                action="failed", error=str(exc),
            ))
            continue

    db.commit()

    return ImportSummary(
        total=len(results),
        inserted=inserted,
        updated=updated,
        failed=failed,
        rows=results,
    )


def preview_excel(file_bytes: bytes, limit: int = 20) -> dict:
    """Return headers + first N rows for preview, plus basic column validation."""
    raw_rows = _read_excel(file_bytes)
    if not raw_rows:
        return {"headers": [], "rows": [], "warnings": ["File has no data rows"]}

    headers = list(raw_rows[0].keys())
    missing = REQUIRED_COLS - set(headers)
    warnings = [f"Missing required column: '{c}'" for c in sorted(missing)]

    preview = []
    for i, row in enumerate(raw_rows[:limit], start=2):
        _, err = _validate_row(row, i)
        preview.append({**{k: row.get(k) for k in headers}, "_error": err})

    return {"headers": headers, "rows": preview, "warnings": warnings}
