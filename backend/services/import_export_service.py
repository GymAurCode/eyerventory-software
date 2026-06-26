import csv
import io
import json
import os
import tempfile
from datetime import datetime
from typing import Any

from fastapi import UploadFile
from sqlalchemy.orm import Session

from backend.database import SessionLocal
from backend.models.product import Product
from backend.models.warehouse import Warehouse, WarehouseStock

# ── Import validation ───────────────────────────────────────────────────────

IMPORT_TEMPLATES = {
    "products": {
        "columns": ["name", "sku", "barcode_number", "category", "cost_price", "selling_price", "stock", "low_stock_threshold"],
        "required": ["name", "cost_price"],
    },
    "suppliers": {
        "columns": ["name", "phone", "address", "email", "opening_balance"],
        "required": ["name"],
    },
    "warehouses": {
        "columns": ["name", "code", "location"],
        "required": ["name"],
    },
    "opening_stock": {
        "columns": ["product_sku", "product_name", "warehouse_code", "quantity", "unit_price"],
        "required": ["product_sku", "warehouse_code", "quantity"],
    },
}


def parse_upload(file: UploadFile) -> tuple[list[dict], list[str]]:
    ext = os.path.splitext(file.filename or "")[1].lower()
    content = file.file.read()

    if ext == ".csv":
        return _parse_csv(content)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(content, ext)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _parse_csv(content: bytes) -> tuple[list[dict], list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    columns = reader.fieldnames or []
    return rows, columns


def _parse_excel(content: bytes, ext: str) -> tuple[list[dict], list[str]]:
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl not installed. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    if not ws:
        raise ValueError("No sheets found in workbook")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter, [])]
    rows = []
    for row in rows_iter:
        rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)})

    return rows, headers


ValidationResult = dict[str, Any]


def validate_import_rows(rows: list[dict], template_key: str) -> ValidationResult:
    template = IMPORT_TEMPLATES.get(template_key)
    if not template:
        return {"valid": False, "errors": [f"Unknown template: {template_key}"], "rows": rows}

    errors = []
    valid_rows = []
    required = template["required"]

    for idx, row in enumerate(rows):
        row_errors = []
        for col in required:
            val = row.get(col, "")
            if not val or str(val).strip() == "":
                row_errors.append(f"Row {idx + 2}: '{col}' is required")

        for col, val in row.items():
            if col in ("cost_price", "selling_price", "unit_price", "opening_balance"):
                try:
                    float(val) if val and str(val).strip() else None
                except (ValueError, TypeError):
                    row_errors.append(f"Row {idx + 2}: '{col}' must be a number")

            if col in ("quantity", "stock", "low_stock_threshold"):
                try:
                    int(val) if val and str(val).strip() else None
                except (ValueError, TypeError):
                    row_errors.append(f"Row {idx + 2}: '{col}' must be an integer")

        if row_errors:
            errors.extend(row_errors)
            row["_errors"] = row_errors
            row["_valid"] = False
        else:
            row["_valid"] = True
        valid_rows.append(row)

    return {
        "valid": len(errors) == 0,
        "total": len(rows),
        "valid_count": sum(1 for r in valid_rows if r["_valid"]),
        "error_count": len(errors),
        "errors": errors,
        "rows": valid_rows,
    }


def import_validated_rows(db: Session, template_key: str, rows: list[dict], dry_run: bool = False) -> dict:
    imported = 0
    errors = []

    for idx, row in enumerate(rows):
        if not row.get("_valid", False):
            continue

        try:
            if template_key == "products":
                _import_product(db, row)
            elif template_key == "opening_stock":
                _import_opening_stock(db, row)
            imported += 1
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")

    if dry_run:
        db.rollback()
        return {"dry_run": True, "would_import": imported, "errors": errors}

    if errors:
        db.rollback()
        return {"success": False, "imported": imported, "errors": errors}

    db.commit()
    return {"success": True, "imported": imported, "errors": []}


def _import_product(db: Session, row: dict):
    existing = db.query(Product).filter(
        (Product.sku == row.get("sku")) | (Product.name == row.get("name"))
    ).first()
    if existing:
        existing.cost_price = float(row.get("cost_price", existing.cost_price or 0))
        existing.selling_price = float(row.get("selling_price", existing.selling_price or 0))
        if row.get("barcode_number"):
            existing.barcode_number = row["barcode_number"]
        if row.get("category"):
            existing.category = row["category"]
        if row.get("low_stock_threshold"):
            existing.low_stock_threshold = int(row["low_stock_threshold"])
    else:
        prod = Product(
            name=row["name"],
            sku=row.get("sku"),
            barcode_number=row.get("barcode_number"),
            category=row.get("category"),
            cost_price=float(row.get("cost_price", 0)),
            selling_price=float(row.get("selling_price", 0)),
            stock=0,
            low_stock_threshold=int(row.get("low_stock_threshold", 10)),
        )
        db.add(prod)


def _import_opening_stock(db: Session, row: dict):
    product = db.query(Product).filter(
        (Product.sku == row.get("product_sku")) | (Product.name == row.get("product_name"))
    ).first()
    if not product:
        raise ValueError(f"Product not found: {row.get('product_sku', row.get('product_name', 'unknown'))}")

    warehouse = db.query(Warehouse).filter(
        (Warehouse.code == row.get("warehouse_code")) | (Warehouse.name == row.get("warehouse_code"))
    ).first()
    if not warehouse:
        raise ValueError(f"Warehouse not found: {row.get('warehouse_code')}")

    ws = db.query(WarehouseStock).filter(
        WarehouseStock.warehouse_id == warehouse.id,
        WarehouseStock.product_id == product.id,
    ).first()
    if ws:
        ws.quantity = int(row["quantity"])
    else:
        db.add(WarehouseStock(
            warehouse_id=warehouse.id,
            product_id=product.id,
            quantity=int(row["quantity"]),
        ))

    product.stock = (
        db.query(db.query(WarehouseStock.quantity).filter(WarehouseStock.product_id == product.id).subquery())
        .statement
    )


# ── Export ──────────────────────────────────────────────────────────────────

def export_to_csv(columns: list[str], rows: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(c, "") for c in columns])
    return output.getvalue()


def export_to_excel(columns: list[str], rows: list[dict], sheet_name: str = "Sheet1") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ValueError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    accent = "f6ce3a"
    header_fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.border = thin_border

    ws.freeze_panes = "A2"

    for ci, col in enumerate(columns, 1):
        max_len = max(
            len(str(col)),
            max((len(str(row.get(col, ""))) for row in rows), default=0),
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def get_products_export_rows(db: Session) -> tuple[list[str], list[dict]]:
    columns = ["name", "sku", "barcode_number", "category", "cost_price", "selling_price", "stock", "low_stock_threshold"]
    products = db.query(Product).all()
    rows = [
        {
            "name": p.name, "sku": p.sku or "", "barcode_number": p.barcode_number or "",
            "category": p.category or "", "cost_price": p.cost_price, "selling_price": p.selling_price,
            "stock": p.stock, "low_stock_threshold": p.low_stock_threshold,
        }
        for p in products
    ]
    return columns, rows


def get_warehouse_export_rows(db: Session) -> tuple[list[str], list[dict]]:
    columns = ["warehouse", "product_name", "sku", "quantity", "reorder_level"]
    stocks = db.query(WarehouseStock).join(Product).join(Warehouse).all()
    rows = [
        {
            "warehouse": s.warehouse.name,
            "product_name": s.product.name,
            "sku": s.product.sku or "",
            "quantity": s.quantity,
            "reorder_level": s.reorder_level,
        }
        for s in stocks
    ]
    return columns, rows
