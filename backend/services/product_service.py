import io
import logging
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from backend.models.product import Product
from backend.schemas.product import BulkImportResult, ProductAddStock, ProductCreate, ProductUpdate

logger = logging.getLogger("inventory-products")

# Required columns for bulk import — must match exactly (case-insensitive strip)
REQUIRED_COLUMNS = {"product_name", "sku", "category", "purchase_price", "sale_price", "stock_quantity"}


def _validate_image(image_data: Optional[str], image_mime: Optional[str]) -> Optional[str]:
    if not image_data:
        return None
    mime = (image_mime or "").strip().lower()
    if mime not in {"image/jpeg", "image/jpg", "image/png"}:
        raise ValueError("Only JPG and PNG images are allowed")
    if not image_data.startswith("data:image/"):
        raise ValueError("Invalid image payload")
    return mime


def list_products(db: Session):
    return db.query(Product).order_by(Product.id.desc()).all()


def create_product(db: Session, payload: ProductCreate):
    # Check for duplicates before creating
    from backend.services.ai_intelligence_service import AIIntelligenceService
    ai_service = AIIntelligenceService(db)
    duplicate_check = ai_service.check_duplicate_risk(payload.name, payload.sku)
    
    # If exact duplicate, raise error
    if duplicate_check.get("exact_duplicate"):
        raise ValueError(f"Product '{payload.name}' already exists")
    
    # If high similarity risk, we'll still create but log warning
    if duplicate_check.get("is_duplicate_risk"):
        logger.warning(
            f"Creating product '{payload.name}' with similar products: "
            f"{[p['name'] for p in duplicate_check.get('similar_products', [])]}"
        )
    
    mime = _validate_image(payload.image_data, payload.image_mime)
    product = Product(
        name=payload.name,
        sku=payload.sku or None,
        category=payload.category or None,
        cost_price=payload.cost_price,
        selling_price=payload.selling_price,
        stock=payload.stock,
        image_data=payload.image_data,
        image_mime=mime,
    )
    db.add(product)
    db.commit()
    db.refresh(product)
    return product


def update_product(db: Session, product_id: int, payload: ProductUpdate):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return None
    updates = payload.model_dump(exclude_none=True)
    if "image_data" in updates:
        updates["image_mime"] = _validate_image(updates.get("image_data"), updates.get("image_mime"))
    for key, value in updates.items():
        setattr(product, key, value)
    db.commit()
    db.refresh(product)
    return product


def delete_product(db: Session, product_id: int) -> bool:
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return False
    db.delete(product)
    db.commit()
    return True


def add_stock(db: Session, product_id: int, payload: ProductAddStock):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return None
    product.stock += payload.quantity
    if payload.price is not None:
        product.cost_price = payload.price
    db.commit()
    db.refresh(product)
    return product


# ── Bulk import ───────────────────────────────────────────────────────────────

def bulk_import_products(db: Session, file_bytes: bytes) -> BulkImportResult:
    """
    Parse an Excel file and upsert products.
    - SKU match → UPDATE
    - No SKU match → CREATE
    - All rows processed in a single transaction; any DB error rolls back everything.
    """
    # 1. Parse workbook
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Cannot read file. Please upload a valid .xlsx file.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The uploaded file is empty.")

    # 2. Validate header row
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    missing = REQUIRED_COLUMNS - set(header)
    if missing:
        raise ValueError(
            f"Invalid file format. Please use the provided template. "
            f"Missing columns: {', '.join(sorted(missing))}"
        )

    col = {name: header.index(name) for name in REQUIRED_COLUMNS}

    # 3. Process data rows
    created = updated = failed = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip fully empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_sku = str(row[col["sku"]] or "").strip() or None
        row_name = str(row[col["product_name"]] or "").strip()

        try:
            # Validate required fields
            if not row_name:
                raise ValueError("product_name is required")

            purchase_price = float(row[col["purchase_price"]] or 0)
            sale_price = float(row[col["sale_price"]] or 0)
            stock_qty = int(row[col["stock_quantity"]] or 0)

            if purchase_price <= 0:
                raise ValueError("purchase_price must be > 0")
            if sale_price < 0:
                raise ValueError("sale_price must be >= 0")
            if stock_qty < 0:
                raise ValueError("stock_quantity must be >= 0")

            category = str(row[col["category"]] or "").strip() or None

            # Upsert logic
            existing = None
            if row_sku:
                existing = db.query(Product).filter(Product.sku == row_sku).first()

            if existing:
                existing.name = row_name
                existing.category = category
                existing.cost_price = purchase_price
                existing.selling_price = sale_price
                existing.stock = stock_qty
                updated += 1
            else:
                db.add(Product(
                    name=row_name,
                    sku=row_sku,
                    category=category,
                    cost_price=purchase_price,
                    selling_price=sale_price,
                    stock=stock_qty,
                ))
                created += 1

        except Exception as exc:
            failed += 1
            errors.append({"row": row_idx, "sku": row_sku or "—", "reason": str(exc)})

    if failed == len(rows) - 1 and failed > 0:
        raise ValueError("All rows failed validation. Please check the file and try again.")

    # 4. Commit everything in one transaction
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise ValueError(f"Database error during import: {exc}") from exc

    return BulkImportResult(
        total=created + updated + failed,
        created=created,
        updated=updated,
        failed=failed,
        errors=errors,
    )


def generate_template_bytes() -> bytes:
    """Return a minimal .xlsx template with the required headers and one example row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["product_name", "sku", "category", "purchase_price", "sale_price", "stock_quantity"]
    ws.append(headers)
    ws.append(["Example Product", "SKU-001", "Electronics", 500.00, 750.00, 10])

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4F52A3", end_color="4F52A3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
